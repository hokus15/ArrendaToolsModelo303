#!/usr/bin/env python3
r"""Generate a schema YAML file from an AEAT Excel specification.

Usage
-----
    python tools/generate_schema.py <excel_path> \
        [--year YYYY] [--revision X.XX] [--output path/to/schema.yaml] \
        [--spec-date YYYY-MM-DD]

Example
-------
    python tools/generate_schema.py "specs/2026/20260128 - DR303e26v101.xlsx" \
        --year 2026 --revision 1.01 \
        --output src/arrendatools/modelo303/infrastructure/schemas/2026.1.yaml

"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

# Ensure src/ is on the path when running from repo root.
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "src"))

try:
    import openpyxl  # noqa: F401 — checked early to give a clear error
except ImportError:
    print(
        "ERROR: openpyxl is required to generate schemas from Excel files.\n"
        "Install it with: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

import yaml  # noqa: E402

from arrendatools.modelo303.infrastructure.builtins import BUILTIN_REGISTRY  # noqa: E402

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Sheets to import (those starting with "DP303" or named like REGISTRO_*)
VALID_SHEET_PREFIXES = ("DP303", "REGISTRO")

# Column aliases → canonical name
COLUMN_ALIASES: dict[str, str] = {
    "nº": "num",
    "n°": "num",
    "posic.": "position",
    "posición": "position",
    "posicion": "position",
    "lon": "length",
    "longitud": "length",
    "tipo": "field_type",
    "descripción": "description",
    "descripcion": "description",
    "validación": "validation",
    "validacion": "validation",
    "contenido": "content",
    "comp": "comp",
}

# Tipo column → field_type value
TIPO_MAP: dict[str, str] = {
    "a": "alphabetical",
    "an": "alphanumeric",
    "num": "numeric",
    "n": "numeric_signed",
}

# Regex to detect constant content: Constante "VALUE"
_CONSTANT_RE = re.compile(r'constante\s+"([^"]*)"', re.IGNORECASE)

# Regex to extract date from filename: YYYYMMDD
_DATE_IN_FILENAME_RE = re.compile(r"(\d{4})(\d{2})(\d{2})")

# Regex to extract casilla reference numbers like [27] or [151] from descriptions
_CASILLA_REF_RE = re.compile(r"\[(\d+)\]")

# Regex for valid formula inner content: [NN] (op [NN])+ with optional spaces
_FORMULA_INNER_RE = re.compile(r"^\s*\[\d+\](\s*[+\-]\s*\[\d+\])+\s*$")


def _infer_formula(description: str) -> tuple[str, str] | None:
    """Infer an arithmetic formula from the field description.

    If the description encodes a single arithmetic formula such as
    '... ( [A] + [B] - [C] ) [result] ...', returns (result_name, expr_str).

    Returns None when:
    - there are zero or more than one parenthesised group containing [NN] refs
      (conditional formulas like 'Si 1T: (...) [N] / Si 4T: (...) [N]'), or
    - the group content contains anything other than [NN] refs and +/- operators.
    """
    all_paren = list(re.finditer(r"\(([^()]+)\)", description))
    # Keep only groups that contain at least one casilla reference [NN]
    formula_parens = [m for m in all_paren if _CASILLA_REF_RE.search(m.group(1))]
    if len(formula_parens) != 1:
        return None
    inner = formula_parens[0].group(1)
    if not _FORMULA_INNER_RE.match(inner):
        return None
    all_refs = _CASILLA_REF_RE.findall(description)
    if not all_refs:
        return None
    result_name = f"casilla_{int(all_refs[-1]):02d}"
    expr = _CASILLA_REF_RE.sub(
        lambda m: f"casilla_{int(m.group(1)):02d}", inner.strip()
    )
    return result_name, expr


# Description keyword → model attribute (checked in order, most specific first)
_DESC_TO_ATTR: list[tuple[str, str]] = [
    # Program metadata
    ("nif empresa desarrollo", "developer_nif"),
    ("versión del programa", "version"),
    ("version del programa", "version"),
    # Banking / refund
    ("código bic", "swift"),
    ("codigo bic", "swift"),
    ("swift", "swift"),
    ("iban", "iban"),
    ("sepa", "sepa"),
    ("banco/bank name", "bank_name"),
    ("bank name", "bank_name"),
    ("entidad bancaria", "bank_name"),
    ("dirección del banco", "bank_address"),
    ("direccion del banco", "bank_address"),
    ("bank address", "bank_address"),
    ("domicilio de la entidad", "bank_address"),
    ("ciudad/city", "bank_city"),
    ("city", "bank_city"),
    ("plaza", "bank_city"),
    ("país/country", "bank_country"),
    ("pais/country", "bank_country"),
    ("country", "bank_country"),
    ("país de la entidad", "bank_country"),
    # Identification flags — most specific first
    ("revocación de la opción por la aplicación de la prorrata", "revocacion_prorrata"),
    ("revocacion de la opcion por la aplicacion de la prorrata", "revocacion_prorrata"),
    ("opción por la aplicación de la prorrata especial", "prorrata_especial"),
    ("opcion por la aplicacion de la prorrata especial", "prorrata_especial"),
    (
        "destinatario de operaciones acogidas al régimen especial del criterio de caja",
        "destinatario_caja",
    ),
    (
        "destinatario de operaciones acogidas al regimen especial del criterio de caja",
        "destinatario_caja",
    ),
    ("criterio de caja", "criterio_caja"),
    ("autoliquidación conjunta", "autoliquidacion_conjunta"),
    ("autoliquidacion conjunta", "autoliquidacion_conjunta"),
    ("tipo de autoliquidación", "tipo_auto_concurso"),
    ("tipo de autoliquidacion", "tipo_auto_concurso"),
    ("fecha en que se dictó el auto", "fecha_concurso"),
    ("fecha en que se dicto el auto", "fecha_concurso"),
    ("concurso de acreedores", "concurso"),
    ("exonerado de la declaración-resumen anual del iva", "exencion_390"),
    ("exonerado de la declaracion-resumen anual del iva", "exencion_390"),
    ("modelo 390", "exencion_390"),
    ("volumen anual de operaciones distinto de cero", "operaciones_no_cero"),
    ("acogido voluntariamente al sii", "sii"),
    ("tributa exclusivamente en régimen simplificado", "solo_rg"),
    ("tributa exclusivamente en regimen simplificado", "solo_rg"),
    ("registro de devolución mensual", "registro_devolucion"),
    ("registro de devolucion mensual", "registro_devolucion"),
    ("tributación exclusivamente foral", "tributacion_foral"),
    ("tributacion exclusivamente foral", "tributacion_foral"),
    ("prorrata especial", "prorrata_especial"),
    ("gasolina", "indicador_pago_gasolina"),
    ("sin actividad", "sin_actividad"),
    ("número justificante", "num_justificante"),
    ("numero justificante", "num_justificante"),
    ("num. justificante", "num_justificante"),
    ("número de justificante", "num_justificante"),
    ("numero de justificante", "num_justificante"),
    ("dar de baja", "baja_domiciliacion"),
    ("baja de domiciliación", "baja_domiciliacion"),
    ("baja de domiciliacion", "baja_domiciliacion"),
    ("motivo de la discrepancia", "motivo_discrepancia"),
    ("discrepancia criterio", "motivo_discrepancia"),
    ("motivo de la rectificación", "motivo_rectificacion"),
    ("motivo de la rectificacion", "motivo_rectificacion"),
    ("rectificativa", "rectificativa"),
    ("con terceras personas", "marque_operaciones_terceras"),
    ("operaciones con terceras personas", "marque_operaciones_terceras"),
    # Core identification
    ("tipo declaración", "tipo_declaracion"),
    ("tipo declaracion", "tipo_declaracion"),
    ("apellidos y nombre", "razon_social"),
    ("razón social", "razon_social"),
    ("razon social", "razon_social"),
    ("ejercicio de devengo", "ejercicio"),
    ("devengo (2) - ejercicio", "ejercicio"),
    ("devengo (2) - período", "periodo"),
    ("devengo (2) - periodo", "periodo"),
    ("período. (pp)", "periodo"),
    ("periodo. (pp)", "periodo"),
    ("período (pp)", "periodo"),
    ("periodo (pp)", "periodo"),
    ("indicador de página complementaria", "pagina_complementaria"),
    ("indicador de pagina complementaria", "pagina_complementaria"),
    ("identificación (1) - nif", "nif"),
    ("identificacion (1) - nif", "nif"),
    # Activity codes (specific before generic)
    # Dash-notation variants from AEAT Excel ("Código de actividad - Principal", etc.)
    ("código de actividad - principal", "codigo_actividad_principal"),
    ("codigo de actividad - principal", "codigo_actividad_principal"),
    ("epígrafe iae - principal", "epigrafe_iae_principal"),
    ("epigrafe iae - principal", "epigrafe_iae_principal"),
    ("código de actividad - otras - 1ª", "codigo_actividad_1"),
    ("codigo de actividad - otras - 1ª", "codigo_actividad_1"),
    ("epígrafe iae - otras - 1ª", "epigrafe_iae_1"),
    ("epigrafe iae - otras - 1ª", "epigrafe_iae_1"),
    ("código de actividad - otras - 2ª", "codigo_actividad_2"),
    ("codigo de actividad - otras - 2ª", "codigo_actividad_2"),
    ("epígrafe iae - otras - 2ª", "epigrafe_iae_2"),
    ("epigrafe iae - otras - 2ª", "epigrafe_iae_2"),
    ("código de actividad - otras - 3ª", "codigo_actividad_3"),
    ("codigo de actividad - otras - 3ª", "codigo_actividad_3"),
    ("epígrafe iae - otras - 3ª", "epigrafe_iae_3"),
    ("epigrafe iae - otras - 3ª", "epigrafe_iae_3"),
    ("código de actividad - otras - 4ª", "codigo_actividad_4"),
    ("codigo de actividad - otras - 4ª", "codigo_actividad_4"),
    ("epígrafe iae - otras - 4ª", "epigrafe_iae_4"),
    ("epigrafe iae - otras - 4ª", "epigrafe_iae_4"),
    ("código de actividad - otras - 5ª", "codigo_actividad_5"),
    ("codigo de actividad - otras - 5ª", "codigo_actividad_5"),
    ("epígrafe iae - otras - 5ª", "epigrafe_iae_5"),
    ("epigrafe iae - otras - 5ª", "epigrafe_iae_5"),
    # Space-notation variants (legacy/alternative descriptions)
    ("código de actividad principal", "codigo_actividad_principal"),
    ("codigo de actividad principal", "codigo_actividad_principal"),
    ("epígrafe del iae principal", "epigrafe_iae_principal"),
    ("epigrafe del iae principal", "epigrafe_iae_principal"),
    ("código de actividad 1", "codigo_actividad_1"),
    ("codigo de actividad 1", "codigo_actividad_1"),
    ("código de actividad 2", "codigo_actividad_2"),
    ("codigo de actividad 2", "codigo_actividad_2"),
    ("código de actividad 3", "codigo_actividad_3"),
    ("codigo de actividad 3", "codigo_actividad_3"),
    ("código de actividad 4", "codigo_actividad_4"),
    ("codigo de actividad 4", "codigo_actividad_4"),
    ("código de actividad 5", "codigo_actividad_5"),
    ("codigo de actividad 5", "codigo_actividad_5"),
    ("epígrafe del iae 1", "epigrafe_iae_1"),
    ("epigrafe del iae 1", "epigrafe_iae_1"),
    ("epígrafe del iae 2", "epigrafe_iae_2"),
    ("epigrafe del iae 2", "epigrafe_iae_2"),
    ("epígrafe del iae 3", "epigrafe_iae_3"),
    ("epigrafe del iae 3", "epigrafe_iae_3"),
    ("epígrafe del iae 4", "epigrafe_iae_4"),
    ("epigrafe del iae 4", "epigrafe_iae_4"),
    ("epígrafe del iae 5", "epigrafe_iae_5"),
    ("epigrafe del iae 5", "epigrafe_iae_5"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _normalise_header(cell_value: object) -> str:
    if cell_value is None:
        return ""
    return str(cell_value).strip().lower()


def _detect_spec_date(excel_path: Path) -> str | None:
    m = _DATE_IN_FILENAME_RE.search(excel_path.stem)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def _map_tipo(raw: str) -> tuple[str, list[str]]:
    """Return (field_type_str, warnings)."""
    key = raw.strip().lower()
    if key in TIPO_MAP:
        return TIPO_MAP[key], []
    # Try without whitespace
    key_nows = re.sub(r"\s+", "", key)
    if key_nows in TIPO_MAP:
        return TIPO_MAP[key_nows], []
    return "alphanumeric", [
        f"Unknown Tipo value {raw!r} -> defaulting to 'alphanumeric'"
    ]


def _analyse_content(content: str) -> tuple[str, str | None, list[str]]:
    """Analyse the Contenido column and return (source, value, warnings)."""
    stripped = content.strip()
    if not stripped:
        return "model", None, []
    lower = stripped.lower()

    # Constante "VALUE"  or  just  "VALUE"  (quoted literal)
    m = _CONSTANT_RE.match(stripped)
    if m:
        return "constant", m.group(1), []
    # "VALUE"  alone or followed by a note reference: "00050". Nota 7.
    if stripped.startswith('"'):
        inner_end = stripped.find('"', 1)
        if inner_end != -1:
            after = stripped[inner_end + 1 :].strip().lstrip(".").strip().lower()  # noqa: E203
            if not after or after.startswith("nota"):
                return "constant", stripped[1:inner_end], []

    # Blanks / spaces
    if lower in ("blancos", "en blanco", "blanco", "rellenar con blancos"):
        return "constant", "", []

    # Ceros
    if lower in ("ceros", "cero", "0"):
        return "default", "0", []

    return "model", None, []


# AEAT requires certain date/indicator fields to render as zeros when blank.
# These fields appear as AN in the Excel specification but must use numeric
# field_type so that the "no value" case encodes as 00000000 (not spaces).
_FIELD_TYPE_OVERRIDES: dict[str, str] = {
    "fecha_concurso": "numeric",
}


def _infer_name(description: str, content: str) -> str | None:
    """
    Infer the model attribute name from the Excel Descripción and Contenido columns.

    Returns a model attribute name string, or None if it cannot be determined.
    """
    # 1. Casilla [NN] references — checked FIRST to avoid false keyword matches.
    #    Spanish tax forms write: "( [A] + [B] ) [result]", so the LAST bracket
    #    reference identifies the result field.
    refs = _CASILLA_REF_RE.findall(description)
    if refs:
        return f"casilla_{int(refs[-1]):02d}"

    # 2. Keyword lookup in _DESC_TO_ATTR (ordered, most-specific first)
    # Collapse all whitespace (including newlines) to single spaces
    desc_lower = " ".join(description.strip().lower().split())
    for keyword, attr in _DESC_TO_ATTR:
        if keyword in desc_lower:
            return attr

    return None


def _find_header_row(sheet, max_scan: int = 20) -> tuple[int | None, dict[str, int]]:
    """Scan rows to find the header row. Returns (row_index, column_map)."""
    for row_idx in range(1, max_scan + 1):
        row_values = [
            _normalise_header(sheet.cell(row=row_idx, column=col).value)
            for col in range(1, sheet.max_column + 1)
        ]
        canon = {COLUMN_ALIASES.get(v, v): idx for idx, v in enumerate(row_values) if v}
        if "position" in canon and "length" in canon:
            return row_idx, canon
    return None, {}


def _cell_str(sheet, row: int, col_map: dict[str, int], key: str) -> str:
    idx = col_map.get(key)
    if idx is None:
        return ""
    val = sheet.cell(row=row, column=idx + 1).value
    return str(val).strip() if val is not None else ""


# ---------------------------------------------------------------------------
# Post-generation field overrides
#
# Some fields cannot be inferred from the Excel spec alone (e.g. casilla_66
# involves a multiplication that _infer_formula does not handle because it
# only recognises + / - expressions).  These overrides are applied after the
# automatic import and keyed by (field_name, source_before_override).
# Each entry replaces the matched field's source/expr/value as specified.
# ---------------------------------------------------------------------------

# Override keyed by field name → dict of keys to set/replace.
# The override is applied to every field whose 'name' matches, regardless of
# which page or year it belongs to.
_FIELD_OVERRIDES: dict[str, dict] = {
    # casilla_66 = casilla_64 * 65 / 100  (percentage of state share)
    # The Excel description says "Atribuible a la Administración del Estado [66]"
    # without a parenthesised arithmetic group, so the auto-importer misses it.
    "casilla_66": {
        "source": "formula",
        "expr": "casilla_64 * casilla_65 / 100",
    },
}


def _apply_field_overrides(schema: dict) -> list[str]:
    """Apply _FIELD_OVERRIDES to matching fields in the schema dict in-place."""
    applied: list[str] = []
    for page in schema.get("pages", []):
        for fld in page.get("fields", []):
            name = fld.get("name", "")
            if name in _FIELD_OVERRIDES:
                overrides = _FIELD_OVERRIDES[name]
                fld.update(overrides)
                # Remove keys that are no longer relevant (e.g. 'value' when
                # switching from constant → formula).
                if overrides.get("source") == "formula":
                    fld.pop("value", None)
                    fld.pop("function", None)
                applied.append(f"[override] {name}: {overrides}")
    return applied


def _parse_field_type(
    raw_tipo: str, sheet_title: str, row_idx: int
) -> tuple[str, list[str]]:
    """Determine field_type from the raw Tipo cell value."""
    if raw_tipo:
        return _map_tipo(raw_tipo)
    return "alphanumeric", [
        f"Sheet '{sheet_title}' row {row_idx}: missing Tipo -> 'alphanumeric'"
    ]


def _apply_desc_override(
    source: str, const_value: str | None, raw_desc: str
) -> tuple[str, str | None]:
    """Override source to 'constant' when the description signals reserved/blank fields."""
    desc_lower = raw_desc.lower()
    if source == "model" and (
        "reservado para la" in desc_lower or "rellenar con blancos" in desc_lower
    ):
        return "constant", ""
    return source, const_value


def _resolve_name_and_source(
    source: str,
    formula_result: tuple[str, str] | None,
    raw_desc: str,
    raw_content: str,
    field_id: str,
    page_id: str,
) -> tuple[str, str, str | None, list[str]]:
    """Resolve the field name, final source, and optional builtin function name."""
    warnings: list[str] = []
    builtin_fn: str | None = None

    if source == "formula":
        field_name = formula_result[0]  # type: ignore[index]
    elif source == "model":
        inferred = _infer_name(raw_desc, raw_content)
        if inferred is not None:
            field_name = inferred
            if field_name in BUILTIN_REGISTRY:
                source = "builtin"
                builtin_fn = field_name
        else:
            field_name = field_id
            warnings.append(
                f"Sheet '{page_id}' field '{field_id}': cannot infer name from "
                f"description {raw_desc!r}; set 'name' manually"
            )
    else:
        inferred = _infer_name(raw_desc, raw_content)
        field_name = inferred if inferred is not None else field_id

    if source == "builtin":
        builtin_fn = field_name

    return field_name, source, builtin_fn, warnings


def _build_field_entry(
    field_id: str,
    field_name: str,
    position: int,
    length: int,
    field_type: str,
    source: str,
    const_value: str | None,
    formula_expr: str | None,
    builtin_fn: str | None,
    raw_tipo: str,
    raw_desc: str,
    raw_val: str,
    raw_content: str,
) -> dict:
    """Assemble the field entry dict."""
    entry: dict = {
        "id": field_id,
        "name": field_name,
        "position": position,
        "length": length,
        "field_type": field_type,
        "source": source,
    }
    if const_value is not None:
        entry["value"] = const_value
    if formula_expr is not None:
        entry["expr"] = formula_expr
    if builtin_fn is not None:
        entry["function"] = builtin_fn
    entry["metadata"] = {
        "tipo": raw_tipo,
        "descripcion": raw_desc,
        "validacion": raw_val,
        "contenido": raw_content,
    }
    return entry


def import_sheet(sheet, page_id: str) -> tuple[list[dict], list[str]]:
    """Parse one Excel sheet into a list of field dicts. Returns (fields, warnings)."""
    header_row, col_map = _find_header_row(sheet)
    if header_row is None:
        return [], [
            f"Sheet '{sheet.title}': could not locate header row (position/length columns)"
        ]

    fields: list[dict] = []
    warnings: list[str] = []
    position = 1
    consecutive_empty = 0

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        raw_len = _cell_str(sheet, row_idx, col_map, "length")
        raw_tipo = _cell_str(sheet, row_idx, col_map, "field_type")
        raw_desc = _cell_str(sheet, row_idx, col_map, "description")
        raw_val = _cell_str(sheet, row_idx, col_map, "validation")
        raw_content = _cell_str(sheet, row_idx, col_map, "content")
        raw_num = _cell_str(sheet, row_idx, col_map, "num")

        # Skip purely decorative or empty rows; bail out after 10 consecutive ones.
        if not raw_len and not raw_tipo and not raw_desc:
            consecutive_empty += 1
            if consecutive_empty >= 10:
                break
            continue
        consecutive_empty = 0
        if raw_num.strip().lower() == "total":
            continue
        if raw_len and not raw_len.replace(" ", "").isdigit():
            continue

        try:
            length = int(raw_len) if raw_len else 0
        except ValueError:
            warnings.append(
                f"Sheet '{sheet.title}' row {row_idx}: cannot parse length {raw_len!r}"
            )
            continue

        if length <= 0:
            continue

        field_type, tipo_warns = _parse_field_type(raw_tipo, sheet.title, row_idx)
        warnings.extend(tipo_warns)

        source, const_value, content_warns = _analyse_content(raw_content)
        warnings.extend(content_warns)

        source, const_value = _apply_desc_override(source, const_value, raw_desc)

        formula_expr: str | None = None
        formula_result = _infer_formula(raw_desc)
        if formula_result is not None:
            source = "formula"
            const_value = None
            formula_expr = formula_result[1]

        try:
            field_number = int(raw_num)
        except (TypeError, ValueError):
            warnings.append(
                f"Sheet '{sheet.title}' row {row_idx}: cannot parse Nº {raw_num!r}"
            )
            continue
        field_id = f"{page_id}_{field_number}"

        field_name, source, builtin_fn, name_warns = _resolve_name_and_source(
            source, formula_result, raw_desc, raw_content, field_id, page_id
        )
        warnings.extend(name_warns)

        if field_name in _FIELD_TYPE_OVERRIDES:
            field_type = _FIELD_TYPE_OVERRIDES[field_name]

        entry = _build_field_entry(
            field_id,
            field_name,
            position,
            length,
            field_type,
            source,
            const_value,
            formula_expr,
            builtin_fn,
            raw_tipo,
            raw_desc,
            raw_val,
            raw_content,
        )
        fields.append(entry)
        position += length

    return fields, warnings


def import_excel(
    excel_path: Path,
    *,
    year: int | None,
    revision: str,
    spec_date: str,
    source_file: str,
) -> tuple[dict, list[str]]:
    """Read the AEAT Excel file and return (schema_dict, all_warnings)."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    all_warnings: list[str] = []
    pages: list[dict] = []
    _registro_cierre: dict | None = None

    for sheet_name in wb.sheetnames:
        upper = sheet_name.strip().upper()
        if not any(upper.startswith(p) for p in VALID_SHEET_PREFIXES):
            continue

        sheet = wb[sheet_name]
        page_id = sheet_name.strip().lower()

        if upper == "DP30300":
            # DP30300 encodes three logical pages:
            #   registro_apertura  → fields at positions 1–17  (num 1-6: <T…0000>)
            #   dp30300            → fields at positions 18–328 (num 7-13: <AUX>…</AUX>)
            #   registro_cierre    → hardcoded at the END of all pages
            all_fields, warns = import_sheet(sheet, page_id)
            all_warnings.extend(warns)

            apertura_fields_raw = [f for f in all_fields if f["position"] <= 17]
            # Rename dp30300_N ids to registro_apertura_N for the opening record
            apertura_fields = []
            for f in apertura_fields_raw:
                fc = dict(f)
                old_id = fc["id"]
                new_id = old_id.replace(f"{page_id}_", "registro_apertura_", 1)
                fc["id"] = new_id
                if fc["name"] == old_id:
                    fc["name"] = new_id
                apertura_fields.append(fc)

            # Re-number positions relative to 1 for the AUX block
            aux_fields_raw = [f for f in all_fields if 18 <= f["position"] <= 328]
            aux_fields = []
            aux_pos = 1
            for f in aux_fields_raw:
                fc = dict(f)
                fc["position"] = aux_pos
                aux_pos += fc["length"]
                aux_fields.append(fc)

            pages.append(
                {
                    "id": "registro_apertura",
                    "include_when": "always",
                    "fields": apertura_fields,
                }
            )
            pages.append(
                {"id": "dp30300", "include_when": "always", "fields": aux_fields}
            )
            # registro_cierre is appended last — stored for deferred append
            _registro_cierre = {
                "id": "registro_cierre",
                "include_when": "always",
                "fields": [
                    {
                        "id": "registro_cierre_1",
                        "name": "registro_general_close",
                        "position": 1,
                        "length": 3,
                        "field_type": "alphanumeric",
                        "source": "constant",
                        "value": "</T",
                        "metadata": {
                            "tipo": "",
                            "descripcion": "",
                            "validacion": "",
                            "contenido": "",
                        },
                    },
                    {
                        "id": "registro_cierre_2",
                        "name": "modelo",
                        "position": 4,
                        "length": 3,
                        "field_type": "numeric",
                        "source": "constant",
                        "value": "303",
                        "metadata": {
                            "tipo": "",
                            "descripcion": "",
                            "validacion": "",
                            "contenido": "",
                        },
                    },
                    {
                        "id": "registro_cierre_3",
                        "name": "discriminante",
                        "position": 7,
                        "length": 1,
                        "field_type": "numeric",
                        "source": "constant",
                        "value": "0",
                        "metadata": {
                            "tipo": "",
                            "descripcion": "",
                            "validacion": "",
                            "contenido": "",
                        },
                    },
                    {
                        "id": "registro_cierre_4",
                        "name": "ejercicio",
                        "position": 8,
                        "length": 4,
                        "field_type": "numeric",
                        "source": "model",
                        "metadata": {
                            "tipo": "",
                            "descripcion": "",
                            "validacion": "",
                            "contenido": "",
                        },
                    },
                    {
                        "id": "registro_cierre_5",
                        "name": "periodo",
                        "position": 12,
                        "length": 2,
                        "field_type": "alphanumeric",
                        "source": "model",
                        "metadata": {
                            "tipo": "",
                            "descripcion": "",
                            "validacion": "",
                            "contenido": "",
                        },
                    },
                    {
                        "id": "registro_cierre_6",
                        "name": "tipo_y_cierre",
                        "position": 14,
                        "length": 5,
                        "field_type": "alphanumeric",
                        "source": "constant",
                        "value": "0000>",
                        "metadata": {
                            "tipo": "",
                            "descripcion": "",
                            "validacion": "",
                            "contenido": "",
                        },
                    },
                ],
            }
            continue

        # DP30302 is the régimen simplificado (RS) page; it does not apply to
        # arrendadores de locales (RG taxpayers) and must never be included.
        if upper == "DP30302":
            continue

        fields_list, warns = import_sheet(sheet, page_id)
        all_warnings.extend(warns)

        # Determine include_when from page name
        if "304" in page_id or "305" in page_id:
            include_when = "fourth_quarter"
        else:
            include_when = "always"

        pages.append(
            {
                "id": page_id,
                "include_when": include_when,
                "fields": fields_list,
            }
        )

    # registro_cierre must appear last (after all declaration pages)
    if _registro_cierre is not None:
        pages.append(_registro_cierre)

    schema_id = f"modelo303-{year}.1" if year else "modelo303-unknown"
    schema: dict = {
        "schema_id": schema_id,
        "fiscal_year": year or 0,
        "revision": revision,
        "specification_date": spec_date,
        "source_file": source_file,
        "pages": pages,
    }

    override_notes = _apply_field_overrides(schema)
    if override_notes:
        all_warnings.extend(override_notes)

    return schema, all_warnings


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Generate a schema YAML file from an AEAT Excel Modelo 303 specification."
    )
    p.add_argument("excel", type=Path, help="Path to the AEAT Excel file.")
    p.add_argument("--year", type=int, help="Fiscal year (e.g. 2026).")
    p.add_argument("--revision", default="1.01", help="Spec revision (default: 1.01).")
    p.add_argument("--output", type=Path, help="Output YAML path.")
    p.add_argument(
        "--spec-date",
        dest="spec_date",
        help="Specification date YYYY-MM-DD. Auto-detected from filename if omitted.",
    )
    return p


def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    excel_path: Path = args.excel.resolve()
    if not excel_path.exists():
        print(f"ERROR: File not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    spec_date: str | None = args.spec_date or _detect_spec_date(excel_path)
    if not spec_date:
        print(
            "ERROR: Could not auto-detect specification date from filename.\n"
            "Pass it explicitly with --spec-date YYYY-MM-DD.",
            file=sys.stderr,
        )
        sys.exit(1)

    output_path: Path = args.output or (
        ROOT
        / "src"
        / "arrendatools"
        / "modelo303"
        / "infrastructure"
        / "schemas"
        / f"{args.year or 'unknown'}.1.yaml"
    )

    source_file = (
        str(excel_path.relative_to(ROOT))
        if excel_path.is_relative_to(ROOT)
        else str(excel_path)
    )

    print(f"Generating schema from {excel_path.name} …")
    schema, warnings = import_excel(
        excel_path,
        year=args.year,
        revision=args.revision,
        spec_date=spec_date,
        source_file=source_file,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fh:
        yaml.dump(
            schema,
            fh,
            allow_unicode=True,
            sort_keys=False,
            default_flow_style=False,
            width=sys.maxsize,  # prevent line-folding that produces YAML list artifacts
        )

    # Summary
    pages = schema["pages"]
    all_fields = [f for p in pages for f in p["fields"]]
    total_fields = len(all_fields)
    by_source: dict[str, int] = {}
    for f in all_fields:
        s = f.get("source", "?")
        by_source[s] = by_source.get(s, 0) + 1

    print(f"\nSchema generated: {output_path}")
    print(f"  Pages:   {len(pages)}")
    print(f"  Fields:  {total_fields}")
    for src in ("constant", "model", "builtin", "formula", "default"):
        count = by_source.get(src, 0)
        if count:
            print(f"    {src:<10} {count}")

    # Separate overrides (informational) from real warnings (need attention)
    overrides = [w for w in warnings if w.startswith("[override]")]
    real_warnings = [w for w in warnings if not w.startswith("[override]")]

    if overrides:
        print(f"\nOverrides applied ({len(overrides)}):")
        for o in overrides:
            print(f"  {o}")

    if real_warnings:
        print(f"\nWarnings — requires manual review ({len(real_warnings)}):")
        for w in real_warnings:
            print(f"  ! {w}")
    else:
        print("\nOK — no manual review needed.")


# Entry point
if __name__ == "__main__":
    main()
if __name__ == "__main__":
    main()
